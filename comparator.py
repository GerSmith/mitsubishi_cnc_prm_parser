# comparator.py

from models import PrmFile, Parameter
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def compare_prm_files(
    old_prm: PrmFile,
    new_prm: PrmFile,
    output_path: Path,
    descriptions: dict = None,
    use_axis_names: bool = False,
    file1_name: str = "OLD",
    file2_name: str = "NEW",
):
    old_params = old_prm.parameters
    new_params = new_prm.parameters

    # Алиасы осей (если нужно)
    axis_names_map = _get_axis_names(new_prm) if use_axis_names else {}

    all_keys = set(old_params.keys()) | set(new_params.keys())
    diff_rows = []

    for key in sorted(all_keys):
        old_param = old_params.get(key)
        new_param = new_params.get(key)

        if old_param is None:
            # Added
            diff_rows.append(_make_row(new_param, None, "Added", descriptions, axis_names_map))
        elif new_param is None:
            # Removed
            diff_rows.append(_make_row(old_param, None, "Removed", descriptions, axis_names_map))
        elif old_param.value != new_param.value:
            # Modified
            diff_rows.append(
                _make_row(new_param, old_param.value, "Modified", descriptions, axis_names_map)
            )

    # Экспорт в Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Diff"

    headers = [
        "Parameter",
        "Axis",
        "Tool",
        "Keep",
        f"Value ({file1_name})",
        f"Value ({file2_name})",
        "Changed",
        "Description",
    ]
    ws.append(headers)

    # Стили
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row_data in diff_rows:
        row = ws.append(row_data)
        last_row = ws.max_row
        change_type = row_data[6]  # Changed
        if change_type == "Added":
            for cell in ws[last_row]:
                cell.fill = green_fill
        elif change_type == "Removed":
            for cell in ws[last_row]:
                cell.fill = red_fill

    # Закрепить шапку
    ws.freeze_panes = "A2"

    wb.save(output_path)


def _get_axis_names(prm: PrmFile) -> dict:
    axis_names = {}
    for axis_num in range(1, 10):
        key = f"1013_A{axis_num}"
        if key in prm.parameters:
            alias = prm.parameters[key].value
            if alias:
                axis_names[axis_num] = alias
            else:
                axis_names[axis_num] = f"A{axis_num}"
        else:
            axis_names[axis_num] = f"A{axis_num}"
    return axis_names


def _make_row(
    param: Parameter, old_value: str, change_type: str, descriptions: dict, axis_names_map: dict
):
    axis_display = axis_names_map.get(param.axis, param.axis) if param.axis is not None else ""
    desc = descriptions.get(param.number, {}).get("description", "") if descriptions else ""
    return [
        param.number,
        axis_display,
        param.tool if param.tool is not None else "",
        param.keep if param.keep is not None else "",
        old_value if old_value is not None else "",
        param.value,
        change_type,
        desc,
    ]
