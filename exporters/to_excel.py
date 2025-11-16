# exporters/to_excel.py

from pathlib import Path
from openpyxl import Workbook
from models import PrmFile


def _get_axis_names(prm: PrmFile) -> dict:
    """
    Возвращает словарь: {1: 'X', 2: 'Y', ...}
    На основе параметра 1013Ax из файла ALL.PRM.
    Если значение пустое или отсутствует — возвращается оригинальный номер оси.
    """
    axis_names = {}
    for axis_num in range(1, 10):  # A1..A9
        key = f"1013_A{axis_num}"
        if key in prm.parameters:
            alias = prm.parameters[key].value
            if alias:  # если не пусто
                axis_names[axis_num] = alias
            else:
                axis_names[axis_num] = f"A{axis_num}"
        else:
            axis_names[axis_num] = f"A{axis_num}"
    return axis_names


def export_to_excel(
    prm: PrmFile, output_path: Path, descriptions: dict = None, use_axis_names: bool = False
):
    wb = Workbook()

    # Лист 1: Метаданные
    ws_header = wb.active
    ws_header.title = "Header"
    ws_header.append(["Field", "Value"])
    for i, line in enumerate(prm.header.raw_lines, start=1):
        ws_header.append([f"Line {i}", line.lstrip(";")])

    # Получаем название осей, если нужно
    axis_names_map = _get_axis_names(prm) if use_axis_names else {}

    # Вспомогательная функция для отображения оси
    def format_axis(axis):
        if axis is None:
            return ""
        return axis_names_map.get(axis, str(axis))  # если нет — просто число

    # Лист 2: Все параметры
    ws_all = wb.create_sheet(title="Parameters.All")
    headers = [
        "Parameter",
        "Axis",
        "Tool",
        "Keep",
        "Value",
        "Group",
        "Subgroup",
        "Short Name",
        "Description",
    ]
    ws_all.append(headers)

    for key in prm.parameters.keys():
        p = prm.parameters[key]
        desc = descriptions.get(p.number, {}) if descriptions else {}
        ws_all.append(
            [
                p.number,
                format_axis(p.axis),
                p.tool if p.tool is not None else "",
                p.keep if p.keep is not None else "",
                p.value,
                desc.get("group", ""),
                desc.get("subgroup", ""),
                desc.get("shortname", ""),
                desc.get("description", ""),
            ]
        )
    ws_all.freeze_panes = "A2"

    # --- Дополнительно: листы по категориям ---
    def _add_sheet(ws_name: str, condition):
        ws = wb.create_sheet(title=ws_name)
        ws.append(headers)
        for key in prm.parameters.keys():
            p = prm.parameters[key]
            if not condition(p):
                continue
            desc = descriptions.get(p.number, {}) if descriptions else {}
            ws.append(
                [
                    p.number,
                    format_axis(p.axis),
                    p.tool if p.tool is not None else "",
                    p.keep if p.keep is not None else "",
                    p.value,
                    desc.get("group", ""),
                    desc.get("subgroup", ""),
                    desc.get("shortname", ""),
                    desc.get("description", ""),
                ]
            )
        ws.freeze_panes = "A2"

    _add_sheet("Parameters.General", lambda p: p.axis is None and p.tool is None and p.keep is None)
    _add_sheet("Parameters.Axis", lambda p: p.axis is not None)
    _add_sheet("Parameters.Tool", lambda p: p.tool is not None)
    _add_sheet("Parameters.Keep", lambda p: p.keep is not None)

    wb.save(output_path)
